#!/usr/bin/env python
#Graphmodule
from openpyxl import *
from Settings import *
import time
import datetime
ScriptName =  __file__
#-------------------------------- START OF PROGRAM -----------------------------
CurrentTimeClass = None # Needed to validate use of the function in "Save"
def GetTime(): #CurrentTimeClass.XXX
	global CurrentTimeClass
	class CurrentTimeClass:
		Year = time.strftime("%Y")
		Month = time.strftime("%m")
		Day = time.strftime("%d")
		Hour = time.strftime("%H")
		Minutes = time.strftime("%M")
		Seconds = time.strftime("%S")
		CurrentTime = Year+Month+Day+Hour+Minutes+Seconds
		YYMMDDString = time.strftime("%Y-%m-%d")
		WeekdayNR = datetime.datetime.today().weekday()
	return CurrentTimeClass

class ExcelFile:
	class WorkSheet:
		def __init__(self, WorkBookObject, Name):
			self.Name = Name
			self = WorkBookObject.wb.create_sheet(title=Name)
			WorkBookObject.ListOfWorkSheets.append(self)

	def __init__(self, Name):
		self.wb = Workbook()
		self.Name = Name
		self.ListOfWorkSheets = []
		self.ListOfWorkSheets.append(self.wb.active)

	def AddDataToCell(self, Sheet, Column, Row, Data):
		try: 
			Column = GetColumnFromIndex(Column)
		except:
			None
		Sheet[str(Column)+str(Row)] = Data

	def Save(self, PathAndFolder):
		print "Saving", self, "as", self.Name, "in", PathAndFolder
		#ADD CHECK FOR SAME NAME TO NOT OVERWRITE!
		CurrentTimeClass = GetTime()
		Date = CurrentTimeClass.Year+"-"+CurrentTimeClass.Month+"-"+CurrentTimeClass.Day+" "+CurrentTimeClass.Hour+CurrentTimeClass.Minutes
		
		self.FileName = self.Name+" "+Date+".xlsx"
		self.wb.save(PathAndFolder+"/"+self.FileName)

#READ
def GetWorksheet(FilePath, FileName, sheetnr):
	print FilePath
	print FileName
	try:
		print "A"
		wb = load_workbook(FilePath+FileName)
	except:
		print "B"
		try: 
			print "C"
			wb = load_workbook(FilePath+FileName+".xlsx")
		except:
			print "D"
			try: 
				print "E"
				wb = load_workbook(FilePath+FileName+".xlsm")
			except:
				print "Could not find the file, check extension?"
	print "X"
	sheet_name = wb.get_sheet_names()[sheetnr]
	worksheet = wb.get_sheet_by_name(sheet_name)
	return worksheet
def GetCellValue(ColNr, RowNr, Worksheet):
	return Worksheet.cell(row=RowNr, column=ColNr).value
def GetColumnValues_AsList(Worksheet, Col, StartRow, EndRow):
	List = []
	for x in range(0, (EndRow-StartRow)):
		value = Worksheet.cell(row=StartRow+x, column=Col).value
		try:
			List.append(float(value))
		except TypeError:
			print StartRow+x-1
			return List
	return List
def GetColumnFromIndex(Index):
	return utils.get_column_letter(Index)
#WRITE
def AddVerticalData(File, Sheet, Data, DataName, FirstColumn, FirstRow):
	File.AddDataToCell(Sheet, FirstColumn, FirstRow, DataName)
	counter = 1
	for d in Data:
		File.AddDataToCell(Sheet, FirstColumn, str(int(FirstRow)+counter), d)
		counter +=1

def AddHorizontalData(File, Sheet, Data, DataName, FirstColumn, FirstRow):
	File.AddDataToCell(Sheet, GetColumnFromIndex(FirstColumn), FirstRow, DataName)
	counter = 1
	for d in Data:
		File.AddDataToCell(Sheet, GetColumnFromIndex(FirstColumn+counter), FirstRow, d)
		counter +=1
#-------------------------------- END OF PROGRAM -----------------------------
print "'"+ScriptName+"'- module has been loaded."
